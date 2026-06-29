import pandas as pd
from datetime import datetime
from openpyxl import load_workbook


def calcular_diferencas(df_financeiro: pd.DataFrame, df_contabilidade: pd.DataFrame, 
                       salvar_arquivo: bool = True, caminho_saida: str = None) -> dict:
    """
    Calcula diferencas entre Financeiro e Contabilidade (em ambas direcoes).
    
    Parametros:
    -----------
    df_financeiro : pd.DataFrame
        DataFrame normalizado do financeiro (colunas: codigo, cliente, valor)
    df_contabilidade : pd.DataFrame
        DataFrame normalizado da contabilidade (colunas: codigo, cliente, valor)
    salvar_arquivo : bool
        Se True, salva resultado em arquivo Excel
    caminho_saida : str
        Caminho para salvar o arquivo (padrao: diferencas_YYYYMMDD_HHMMSS.xlsx)
    
    Retorna:
    --------
    dict contendo:
        - 'df_completo': DataFrame com todas as diferencas
        - 'resumo': Estatisticas das diferencas
        - 'caminho_arquivo': Caminho do arquivo salvo (se salvar_arquivo=True)
    """
    
    print("[INFO] Calculando diferencas...")
    
    # Padronizar nomes das colunas
    df_fin = df_financeiro.copy()
    df_cont = df_contabilidade.copy()
    
    # Garantir que as colunas existem
    if 'codigo' not in df_fin.columns or 'valor' not in df_fin.columns:
        raise ValueError("df_financeiro deve ter colunas 'codigo' e 'valor'")
    if 'codigo' not in df_cont.columns or 'valor' not in df_cont.columns:
        raise ValueError("df_contabilidade deve ter colunas 'codigo' e 'valor'")
    
    # Fazer merge completo (outer join) para pegar todos os codigos
    df_merge = pd.merge(
        df_fin[['codigo', 'cliente', 'valor']],
        df_cont[['codigo', 'cliente', 'valor']],
        on='codigo',
        how='outer',
        suffixes=('_fin', '_cont')
    )
    
    # Preencher valores NaN com 0
    df_merge['valor_fin'] = df_merge['valor_fin'].fillna(0)
    df_merge['valor_cont'] = df_merge['valor_cont'].fillna(0)
    
    # Usar cliente do financeiro, se nao existir usar da contabilidade
    df_merge['cliente'] = df_merge['cliente_fin'].fillna(df_merge['cliente_cont'])
    
    # Calcular diferenca: Contabilidade - Financeiro
    df_merge['diferenca'] = df_merge['valor_cont'] - df_merge['valor_fin']
    df_merge['diferenca_abs'] = df_merge['diferenca'].abs()
    
    # Calcular percentual de diferenca
    df_merge['diferenca_perc'] = 0.0
    mask = df_merge['valor_fin'] != 0
    df_merge.loc[mask, 'diferenca_perc'] = (
        (df_merge.loc[mask, 'diferenca'] / df_merge.loc[mask, 'valor_fin']) * 100
    )
    
    # Classificar origem dos registros
    df_merge['origem'] = 'Ambos'
    df_merge.loc[df_merge['valor_fin'] == 0, 'origem'] = 'So Contabilidade'
    df_merge.loc[df_merge['valor_cont'] == 0, 'origem'] = 'So Financeiro'
    
    # Classificar tipo de diferenca
    df_merge['tipo_diferenca'] = 'Sem diferenca'
    df_merge.loc[df_merge['diferenca'] > 0, 'tipo_diferenca'] = 'Contabilidade > Financeiro'
    df_merge.loc[df_merge['diferenca'] < 0, 'tipo_diferenca'] = 'Financeiro > Contabilidade'
    df_merge.loc[df_merge['origem'] != 'Ambos', 'tipo_diferenca'] = 'Exclusivo'
    
    # Selecionar e ordenar colunas finais
    df_resultado = df_merge[[
        'codigo',
        'cliente',
        'valor_fin',
        'valor_cont',
        'diferenca',
        'diferenca_abs',
        'diferenca_perc',
        'origem',
        'tipo_diferenca'
    ]].copy()
    
    # Renomear para melhor visualizacao
    df_resultado.columns = [
        'Codigo',
        'Cliente',
        'Valor Financeiro',
        'Valor Contabilidade',
        'Diferenca',
        'Diferenca Absoluta',
        'Diferenca %',
        'Origem',
        'Tipo Diferenca'
    ]
    
    # Ordenar por diferenca absoluta (maiores primeiro)
    df_resultado = df_resultado.sort_values('Diferenca Absoluta', ascending=False)
    
    # Calcular resumo
    resumo = {
        'total_registros': len(df_resultado),
        'registros_ambos': len(df_resultado[df_resultado['Origem'] == 'Ambos']),
        'registros_so_financeiro': len(df_resultado[df_resultado['Origem'] == 'So Financeiro']),
        'registros_so_contabilidade': len(df_resultado[df_resultado['Origem'] == 'So Contabilidade']),
        'registros_com_diferenca': len(df_resultado[df_resultado['Diferenca Absoluta'] > 0.01]),
        'registros_sem_diferenca': len(df_resultado[df_resultado['Diferenca Absoluta'] <= 0.01]),
        'diferenca_total': df_resultado['Diferenca'].sum(),
        'diferenca_absoluta_total': df_resultado['Diferenca Absoluta'].sum(),
        'maior_diferenca': df_resultado['Diferenca Absoluta'].max(),
        'valor_total_financeiro': df_resultado['Valor Financeiro'].sum(),
        'valor_total_contabilidade': df_resultado['Valor Contabilidade'].sum()
    }
    
    print(f"   [OK] Total de registros analisados: {resumo['total_registros']}")
    print(f"   [OK] Registros com diferenca: {resumo['registros_com_diferenca']}")
    print(f"   [OK] Diferenca total: R$ {resumo['diferenca_total']:,.2f}")
    
    # Salvar arquivo se solicitado
    caminho_arquivo = None
    if salvar_arquivo:
        if caminho_saida is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            caminho_arquivo = f"diferencas_{timestamp}.xlsx"
        else:
            caminho_arquivo = caminho_saida
        
        # Criar arquivo Excel com multiplas abas
        print(f"\n[INFO] Salvando arquivo: {caminho_arquivo}")
        
        with pd.ExcelWriter(caminho_arquivo, engine='openpyxl') as writer:
            # Aba 1: Total das diferencas
            df_resultado.to_excel(writer, sheet_name='Total das Diferencas', index=False)
            
            # Aba 2: Apenas com diferencas significativas
            df_com_dif = df_resultado[df_resultado['Diferenca Absoluta'] > 0.01].copy()
            df_com_dif.to_excel(writer, sheet_name='Com Diferencas', index=False)
            
            # Aba 3: Apenas no Financeiro
            df_so_fin = df_resultado[df_resultado['Origem'] == 'So Financeiro'].copy()
            if len(df_so_fin) > 0:
                df_so_fin.to_excel(writer, sheet_name='So Financeiro', index=False)
            
            # Aba 4: Apenas na Contabilidade
            df_so_cont = df_resultado[df_resultado['Origem'] == 'So Contabilidade'].copy()
            if len(df_so_cont) > 0:
                df_so_cont.to_excel(writer, sheet_name='So Contabilidade', index=False)
            
            # Aba 5: Resumo
            df_resumo = pd.DataFrame([resumo]).T
            df_resumo.columns = ['Valor']
            df_resumo.index.name = 'Metrica'
            df_resumo.to_excel(writer, sheet_name='Resumo')
        
        # Aplicar formatacao
        _formatar_arquivo_excel(caminho_arquivo)
        
        print(f"   [OK] Arquivo salvo com {len(df_resultado)} registros")
        print(f"   [OK] Abas criadas: Total das Diferencas, Com Diferencas, So Financeiro, So Contabilidade, Resumo")
    
    return {
        'df_completo': df_resultado,
        'resumo': resumo,
        'caminho_arquivo': caminho_arquivo
    }


def _formatar_arquivo_excel(caminho_arquivo: str):
    """Aplica formatacao ao arquivo Excel gerado"""
    
    wb = load_workbook(caminho_arquivo)
    
    # Formatar cada aba
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        if sheet_name == 'Resumo':
            continue
        
        # Encontrar colunas de valor
        header = [cell.value for cell in ws[1]]
        
        for idx, col_name in enumerate(header, 1):
            if col_name in ['Valor Financeiro', 'Valor Contabilidade', 'Diferenca', 
                           'Diferenca Absoluta']:
                # Aplicar formato de moeda
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=idx)
                    cell.number_format = 'R$ #,##0.00;[RED]-R$ #,##0.00'
            
            elif col_name == 'Diferenca %':
                # Aplicar formato de percentual
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=idx)
                    cell.number_format = '0.00%'
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 12  # Codigo
        ws.column_dimensions['B'].width = 35  # Cliente
        ws.column_dimensions['C'].width = 18  # Valor Financeiro
        ws.column_dimensions['D'].width = 20  # Valor Contabilidade
        ws.column_dimensions['E'].width = 15  # Diferenca
        ws.column_dimensions['F'].width = 18  # Diferenca Absoluta
        ws.column_dimensions['G'].width = 12  # Diferenca %
        ws.column_dimensions['H'].width = 18  # Origem
        ws.column_dimensions['I'].width = 25  # Tipo Diferenca
    
    wb.save(caminho_arquivo)
